from __future__ import annotations

import csv
import shutil
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

INPUT_DIR = Path(__file__).resolve().parent / "in"
OUTPUT_DIR = Path(__file__).resolve().parent / "out"

SHEET_CONFIG = {
    "tempa": {
        "amigavel": "amigavel_A.csv",
        "contencioso": "contencioso_A.csv",
    },
    "tempb": {
        "amigavel": "amigavel_B.csv",
        "contencioso": "contencioso_B.csv",
    },
}


def normalize_sheet_name(name: object) -> str:
    return str(name).strip().lower()


def clear_output_folder(output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)

    for item in output_dir.iterdir():
        if item.name == ".gitkeep" and item.is_file():
            continue

        if item.is_dir():
            shutil.rmtree(item)
        else:
            item.unlink()


def normalize_phone(value: object) -> str | None:
    if value is None:
        return None

    if isinstance(value, int):
        return str(value)

    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else "".join(ch for ch in str(value) if ch.isdigit())

    digits = "".join(ch for ch in str(value).strip() if ch.isdigit())
    return digits or None


def normalize_name(value: object) -> str | None:
    if value is None:
        return None

    name = " ".join(str(value).split())
    if not name:
        return None

    return name.lower().title()


def extract_rows(sheet, phone_col: int, name_col: int) -> list[tuple[str, str]]:
    rows: list[tuple[str, str]] = []

    for row in sheet.iter_rows(min_row=3, values_only=True):
        phone_raw = row[phone_col] if phone_col < len(row) else None
        name_raw = row[name_col] if name_col < len(row) else None

        phone = normalize_phone(phone_raw)
        name = normalize_name(name_raw)

        if phone and name:
            rows.append((phone, name))

    return rows


def write_csv(file_path: Path, data: Iterable[tuple[str, str]]) -> None:
    with file_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["phonenumber", "name"])
        writer.writerows(data)


def process_excel(excel_path: Path, output_dir: Path) -> None:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)

    try:
        sheet_lookup = {normalize_sheet_name(name): name for name in workbook.sheetnames}

        for sheet_name, outputs in SHEET_CONFIG.items():
            actual_sheet_name = sheet_lookup.get(normalize_sheet_name(sheet_name))
            if actual_sheet_name is None:
                available = ", ".join(workbook.sheetnames)
                raise ValueError(
                    f"A planilha '{sheet_name}' nao existe em {excel_path.name}. "
                    f"Planilhas disponiveis: {available}."
                )

            sheet = workbook[actual_sheet_name]

            amigavel_rows = extract_rows(sheet, phone_col=0, name_col=1)
            contencioso_rows = extract_rows(sheet, phone_col=5, name_col=6)

            write_csv(output_dir / outputs["amigavel"], amigavel_rows)
            write_csv(output_dir / outputs["contencioso"], contencioso_rows)
    finally:
        workbook.close()


def find_input_excels(input_dir: Path) -> list[Path]:
    patterns = ("*.xlsx", "*.xlsm", "*.xltx", "*.xltm")
    files: list[Path] = []

    for pattern in patterns:
        files.extend(input_dir.glob(pattern))

    return sorted(files)


def main() -> int:
    INPUT_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    excel_files = find_input_excels(INPUT_DIR)
    if not excel_files:
        print("Nenhum Excel encontrado em base/in.")
        return 1

    clear_output_folder(OUTPUT_DIR)

    for excel_file in excel_files:
        process_excel(excel_file, OUTPUT_DIR)

    for excel_file in excel_files:
        excel_file.unlink(missing_ok=True)

    print("Extracao concluida com sucesso.")
    print("Arquivos gerados:")
    for name in ["amigavel_A.csv", "amigavel_B.csv", "contencioso_A.csv", "contencioso_B.csv"]:
        print(f"- {name}")

    print("Excels removidos de base/in.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
